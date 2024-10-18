from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from subscription.models import Subscription, SubscriptionPlan
from .models import Project, Set, Skill, Duration, Reqskill,MGskill,DS,FS,WAD,Score, Cost, PreferenceCost, Calendar,ProjectFiles
from .forms import SkillForm, DurationForm, SetForm
from django.http import HttpResponse
from django.db.models import F
from collections import defaultdict
import pandas as pd
import random
from django.urls import reverse
import os
import sys
from datetime import timedelta
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter

from ortools.linear_solver import pywraplp
solver = pywraplp.Solver.CreateSolver('SCIP')


# Ensure the static directory for project files exists
staticfiles_dir = os.path.join(settings.BASE_DIR, 'static', 'projectfiles')
if not os.path.exists(staticfiles_dir):
    os.makedirs(staticfiles_dir)

from django.contrib.auth import get_user_model

User = get_user_model()
# Create your views here.
@login_required
def dashboard_view(request):
    # Get the user's subscription
    subscription = Subscription.objects.filter(user=request.user, status='active').first()

    # Get the user's projects and files
    projects = Project.objects.filter(user=request.user)  # Get all projects for the user
    total_projects = projects.count()  # Get the total number of projects
    

    # Pass the subscription, total project count, and files to the template
    return render(request, 'dashboard/index.html', {
        'subscription': subscription,
        'total_projects': total_projects
    })

@login_required
def project_view(request,project_id):
    project = Project.objects.get(id=project_id)
    files = ProjectFiles.objects.filter(user=request.user,project = project)
    
    return render(request, 'dashboard/project.html',{'files':files})
    
@login_required
def calendar_view(request, project_id):
    # Fetch the project based on the provided ID
    project = Project.objects.get(id=project_id)
    
    # Fetch all calendar entries related to the project
    calendars = Calendar.objects.filter(project=project)
    
    # Get unique team leads
    team_leads = calendars.values_list('team_lead', flat=True).distinct()

    # Define the date range (1/1/2024 to 1/6/2024)
    date_range = ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04', '2024-01-05', '2024-01-06']
    
    if request.method == 'POST':
        # Loop over the team leads and dates to update the Calendar values
        for team_lead in team_leads:
            for idx, date in enumerate(date_range):
                # Get the input name dynamically (e.g., "calendar_TL1_0")
                input_name = f'calendar_{team_lead}_{idx}'
                
                # Get the new value from the POST data
                new_value = request.POST.get(input_name)
                
                if new_value is not None:
                    # Find the calendar entry for the specific team lead and date
                    calendar_entry = Calendar.objects.filter(
                        Q(project=project) & Q(team_lead=team_lead) & Q(date=date)
                    ).first()
                    
                    if calendar_entry:
                        # Update the value in the Calendar model
                        calendar_entry.value = int(new_value)
                        calendar_entry.save()

        # Redirect to the same page to prevent form resubmission
        return redirect('calendar_view', project_id=project_id)

    # Prepare the calendar data dictionary
    calendar_data = {}

    for lead in team_leads:
        calendar_data[lead] = []
        for day, date in enumerate(date_range):
            # Get the value for each team lead on each date
            entry = calendars.filter(team_lead=lead, date=date).first()
            if entry:
                calendar_data[lead].append(entry.value)
            else:
                calendar_data[lead].append(0)  # Default to 0 if no entry is found

    context = {
        'project': project,
        'calendar_data': calendar_data,
        'date_range': date_range,
    }

    return render(request, 'dashboard/calendar_view.html', context)

@login_required
def select_project(request):
    if request.method == "POST":
        project_name = request.POST.get('project_name')
        selected_type = request.POST.get('input-style')
        if selected_type == 'software':
            return redirect('create-project')
        else:
            project = Project.objects.create(name=project_name,user = request.user, start_date=None, finish_date=None)
            return redirect('uploadfile', project_id=project.id)
        
@login_required
def uploadfile(request, project_id):
    
    return render(request, 'dashboard/upload.html',{'project_id' :project_id})

@login_required
def create_project(request):
    user = request.user
    if request.method == 'POST':
        project_name = request.POST.get('name')
        start_date = request.POST.get('start_date')
        finish_date = request.POST.get('finish_date')
        project = Project.objects.create(name=project_name,user = user, start_date=start_date, finish_date=finish_date)
        return redirect('create-set', project_id=project.id)
    
    return render(request, 'dashboard/create_project.html')

@login_required
def process(request, project_id):
    # Retrieve the project using get_object_or_404 to handle possible DoesNotExist exception
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        # Handle your form submission here, for example:
        # form = YourForm(request.POST)
        # if form.is_valid():
        #     # Process form data
        #     return HttpResponseRedirect(reverse('some-view'))
        pass  # Placeholder for your POST logic

    # Pass any additional context data here, such as related Sets or other models
    sets = project.sets.all()  # Assuming Set is related to Project

    return render(request, 'dashboard/process.html', {'project': project, 'sets': sets})

@login_required
def create_set(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        # Process form data to create a new Set, converting string values to integers
        projects = int(request.POST.get('projects'))
        teamleads = int(request.POST.get('teamleads'))
        timeperiod = int(request.POST.get('timeperiod'))
        skills = int(request.POST.get('skills'))
        managers = int(request.POST.get('managers'))

        # Save the new set related to the project
        set_instance = Set.objects.create(
            project=project,
            projects=projects,
            teamleads=teamleads,
            timeperiod=timeperiod,
            skills=skills,
            managers=managers
        )

        # Redirect to the page where the user can edit the generated data
        return redirect('edit_project_data', project_id=project.id)

    return render(request, "dashboard/create-set.html", {'project': project})



@login_required
def edit_project_data(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    # Fetch related data filtering by project_id
    durations = Duration.objects.filter(project=project)
    reqskills = Reqskill.objects.filter(project=project)
    mgskills = MGskill.objects.filter(project=project)
    ds_entries = DS.objects.filter(project=project)
    fs_entries = FS.objects.filter(project=project)
    scores = Score.objects.filter(project=project)
    costs = Cost.objects.filter(project=project)
    wads = WAD.objects.filter(project=project)
    calendars = Calendar.objects.filter(project=project)
    preference_costs = PreferenceCost.objects.filter(project=project)

    # Fetch team leads and date range for Calendar data
    team_leads = calendars.values_list('team_lead', flat=True).distinct()
    start_date = project.start_date
    finish_date = project.finish_date
    delta = finish_date - start_date
    date_range = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

    # Fetch project names for PreferenceCost matrix
    project_names = preference_costs.values_list('project_from', flat=True).distinct()

    if request.method == 'POST':
        if 'export_excel' in request.POST:
            # Create a new Excel workbook
            wb = openpyxl.Workbook()

            # Helper function to create sheets and write data for each table
            def create_sheet_with_data(wb, sheet_name, queryset, fields):
                sheet = wb.create_sheet(title=sheet_name)
                # Add header
                for col_num, field in enumerate(fields, 1):
                    col_letter = get_column_letter(col_num)
                    sheet[f'{col_letter}1'] = field

                # Add rows
                for row_num, obj in enumerate(queryset, 2):
                    for col_num, field in enumerate(fields, 1):
                        col_letter = get_column_letter(col_num)
                        sheet[f'{col_letter}{row_num}'] = getattr(obj, field, '')

            # Add data for each model to a separate sheet, filtering by project
            create_sheet_with_data(wb, 'Durations', durations, ['projects', 'value'])
            create_sheet_with_data(wb, 'Reqskills', reqskills, ['projects', 'skills', 'value'])
            create_sheet_with_data(wb, 'MGskills', mgskills, ['skills', 'managers', 'value'])
            create_sheet_with_data(wb, 'DS', ds_entries, ['projects', 'value'])
            create_sheet_with_data(wb, 'FS', fs_entries, ['projects', 'value'])
            create_sheet_with_data(wb, 'Scores', scores, ['teamleads', 'value'])
            create_sheet_with_data(wb, 'Costs', costs, ['teamleads', 'value'])
            create_sheet_with_data(wb, 'WAD', wads, ['teamleads', 'value'])
            create_sheet_with_data(wb, 'Calendars', calendars, ['team_lead', 'date', 'value'])
            create_sheet_with_data(wb, 'PreferenceCosts', preference_costs, ['project_from', 'project_to', 'cost_value'])

            # Remove the default sheet created by openpyxl
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            # Prepare the HttpResponse for Excel file download
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={project.name}_data.xlsx'

            # Save the workbook to the HttpResponse
            wb.save(response)

            return response

        # Update each model's data if form is submitted
        for duration in durations:
            value = request.POST.get(f'duration_{duration.id}')
            if value is not None:
                duration.value = int(value)
                duration.save()

        for reqskill in reqskills:
            value = request.POST.get(f'reqskill_{reqskill.id}')
            if value is not None:
                reqskill.value = int(value)
                reqskill.save()

        for mgskill in mgskills:
            value = request.POST.get(f'mgskill_{mgskill.id}')
            if value is not None:
                mgskill.value = int(value)
                mgskill.save()

        for ds in ds_entries:
            value = request.POST.get(f'ds_{ds.id}')
            if value is not None:
                ds.value = int(value)
                ds.save()

        for fs in fs_entries:
            value = request.POST.get(f'fs_{fs.id}')
            if value is not None:
                fs.value = int(value)
                fs.save()

        for score in scores:
            value = request.POST.get(f'score_{score.id}')
            if value is not None:
                score.value = float(value)
                score.save()

        for cost in costs:
            value = request.POST.get(f'cost_{cost.id}')
            if value is not None:
                cost.value = int(value)
                cost.save()

        for wad in wads:
            value = request.POST.get(f'wad_{wad.id}')
            if value is not None:
                wad.value = int(value)
                wad.save()

        # Save Calendar data
        for lead in team_leads:
            for day, date in enumerate(date_range):
                input_name = f'calendar_{lead}_{day}'
                value = request.POST.get(input_name)

                if value is not None:
                    entry = calendars.filter(team_lead=lead, date=date).first()
                    if entry:
                        entry.value = int(value)
                        entry.save()

        # Save PreferenceCost data
        for row_project in project_names:
            for col_project in project_names:
                input_name = f'pref_cost_{row_project}_{col_project}'
                value = request.POST.get(input_name)

                if value is not None:
                    pref_cost = preference_costs.filter(project_from=row_project, project_to=col_project).first()
                    if pref_cost:
                        pref_cost.cost_value = int(value)
                        pref_cost.save()

        # Redirect after saving
        return redirect('dashboard')

    # Prepare calendar data for rendering
    calendar_data = {}
    for lead in team_leads:
        calendar_data[lead] = []
        for day, date in enumerate(date_range):
            entry = calendars.filter(team_lead=lead, date=date).first()
            if entry:
                calendar_data[lead].append(entry.value)
            else:
                calendar_data[lead].append(0)  # Default to 0 if no entry is found

    # Prepare preference matrix for rendering
    preference_matrix = {}
    for row_project in project_names:
        preference_matrix[row_project] = {}
        for col_project in project_names:
            pref_cost = preference_costs.filter(project_from=row_project, project_to=col_project).first()
            if pref_cost:
                preference_matrix[row_project][col_project] = pref_cost.cost_value
            else:
                preference_matrix[row_project][col_project] = 0  # Default to 0 if no entry is found

    return render(request, 'dashboard/edit-project-data.html', {
        'project': project,
        'durations': durations,
        'reqskills': reqskills,
        'mgskills': mgskills,
        'ds_entries': ds_entries,
        'fs_entries': fs_entries,
        'scores': scores,
        'costs': costs,
        'wads': wads,
        'calendar_data': calendar_data,
        'date_range': date_range,
        'preference_matrix': preference_matrix,
        'project_names': project_names,
    })

from .models import Cost  # Ensure you have this import




def optimize(project_id):
    # Fetch the project and set data
    project = get_object_or_404(Project, id=project_id)
    
    
    set_obj = get_object_or_404(Set, project=project_id)
    
    # Define ranges based on the Set model
    try:
        T = range(set_obj.timeperiod)  # Time periods (e.g., 30 time units)
        S = range(set_obj.skills)      # Skills (e.g., 19 skills)
        W = range(set_obj.teamleads)   # Workforce (e.g., 10 team leads)
        P = range(set_obj.projects)    # Projects (e.g., 20 projects)
    except ValueError as e:
        
        return None

    # Step 1: Create 'Sets' dictionary to store set data
    Sets = {
        'projects': set_obj.projects,
        'teamleads': set_obj.teamleads,
        'timeperiod': set_obj.timeperiod,
        'skills': set_obj.skills,
        'managers': set_obj.managers
    }
    

    # Step 2: Fetch Calendar data
    calendars = Calendar.objects.filter(project=project_id)
    if not calendars.exists():
        
        return None
    
    calendar_df = {}
    for calendar in calendars:
        try:
            lead_index = int(calendar.team_lead.replace('TL', '')) - 1  # Assuming team leads are TL1, TL2, etc.
            date_index = (calendar.date - project.start_date).days  # Number of days since the start date
            calendar_df[(lead_index, date_index)] = calendar.value
        except ValueError as e:
            
            continue
    

    # Step 3: Fetch PreferenceCost (Precedence Matrix)
    preference_costs = PreferenceCost.objects.filter(project=project_id)
    if not preference_costs.exists():
        
        return None
    
    Prec = {}
    for pref_cost in preference_costs:
        try:
            from_index = int(pref_cost.project_from.replace('P', '')) - 1  # Assuming project names are P1, P2, etc.
            to_index = int(pref_cost.project_to.replace('P', '')) - 1
            Prec[(from_index, to_index)] = pref_cost.cost_value
        except ValueError as e:
            
            continue

    print(f"Prec (Prec Availability): {Prec}")
    print(f"Prec keys: {Prec.keys()}")


    # Fetch Required Skills
    reqskills = Reqskill.objects.filter(project=project_id)

    RSkl = {}
    for reqskill in reqskills:
        try:
            # Ensure the 'skills' and 'projects' fields contain valid strings like 'Skill1' and 'P1'
            if 'Skill' in reqskill.skills and 'P' in reqskill.projects:
                skill_index = int(reqskill.skills.replace('Skill', '')) - 1
                project_index = int(reqskill.projects.replace('P', '')) - 1
                RSkl[(skill_index, project_index)] = reqskill.value
            else:
                print(f"Unexpected format in Reqskill record: Skill={reqskill.skills}, Project={reqskill.projects}")
        except ValueError as e:
            print(f"ValueError for Skill={reqskill.skills}, Project={reqskill.projects}: {e}")
            continue  # Skip this record if there's an error
    
    print(f"RSkl (Required Skills Availability): {RSkl}")
    
    reqskills = Reqskill.objects.filter(project=int(project_id))
    for reqskill in reqskills:
        print(f"Project: {reqskill.project}, Skill: {reqskill.skills}, Value: {reqskill.value}")
        

    # Step 5: Fetch Manager Skills Availability
    mgskills = MGskill.objects.filter(project=project_id)
    if not mgskills.exists():
        
        return None
    
    MSAvl = {}
    for mgskill in mgskills:
        try:
            skill_index = int(mgskill.skills.replace('Skill', '')) - 1
            manager_index = int(mgskill.managers.replace('Manager', '')) - 1
            MSAvl[(skill_index, manager_index)] = mgskill.value
        except ValueError as e:
            
            continue

   

    # Step 6: Fetch Duration Data
    durations = Duration.objects.filter(project=project_id)
    if not durations.exists():
        
        return None
    
    Dur = {i: duration.value for i, duration in enumerate(durations)}
    
    # Step 7: Fetch Starting Dates (DS) and Finishing Dates (FS)
    due_start_dates = DS.objects.filter(project=project)
    if not due_start_dates.exists():
        print("error")
        
    
    SD = {i: ds.value for i, ds in enumerate(due_start_dates)}
    

    due_finish_dates = FS.objects.filter(project=project)
    
    
    FD = {i: fs.value for i, fs in enumerate(due_finish_dates)}
    

    # Step 8: Fetch Cost Data
    costs_entries = Cost.objects.filter(project=project_id)
    
    
    Cost_sheet = {i: cost.value for i, cost in enumerate(costs_entries)}
    

    # Step 9: Fetch Workforce Skills Availability
    skills = Skill.objects.filter(project=project_id).all()
    
    
    WSAvl = {}
    
    for skill in skills:
        try:
            # Directly convert the skills and teamleads fields to integers, no need for 'replace'
            skill_index = int(skill.skills) - 1  # Convert to zero-based index
            teamlead_index = int(skill.teamleads.replace('TL', '')) - 1  # Keep only the number from 'TLX'
            
            # Populate WSAvl dictionary with the available skill value for the team lead
            WSAvl[(skill_index, teamlead_index)] = skill.value
        except ValueError as e:
            print(f"Error processing skill {skill.skills} or teamlead {skill.teamleads}: {e}")
            continue  # Skip any entries with errors

    print(f"WSAvl (Worker Skills Availability): {WSAvl}")
        # Step 10: Create the Calendar dictionary from calendar_df
    Cal = {}
    for (i, j), value in calendar_df.items():
        if isinstance(value, (int, float)):
            Cal[(i, j)] = int(value)
    

    # Step 11: Generate Score and WAD values (Workers Available Days)
    WAD = {}
    for worker in W:
        WAD[worker] = sum([Cal.get((worker, t), 0) for t in T])
    

    # Step 12: Compute Score for each worker
    Score = {}
    for w in W:
        score_sum = 0
        for t in T:
            cal_value = Cal.get((w, t), 0)
            score_sum += (len(T) - t + 1 - cal_value) * (1 - cal_value)
        Score[w] = score_sum / len(T)
    
    
    # Fetch all distinct skills for the specific project
    unique_skills = Skill.objects.filter(project=project).values_list('skills', flat=True).distinct()

    # Convert the QuerySet to a list (optional, depending on what you need)
    unique_skills = list(unique_skills)

    print("Start Solving")

    X =  {(p, w): solver.BoolVar(f'X[{p}][{w}]') for p in P for w in W}
    ST = {(p): solver.IntVar(0, 500, f'ST[{p}]') for p in P }
    FT = {(p): solver.IntVar(0, 500, f'FT[{p}]') for p in P }

    # Constraints

    for p in P:
        solver.Add(sum(X[p, w] for w in W) == 1)
        
    for p in P:
        for w in W:
            for s in S:
                # Corrected: WSAvl[s, w] is used instead of [s, w]
                solver.Add(RSkl.get((s, p), 0) - WSAvl.get((s, w), 0) <= 2 * (1 - X[p, w]))

 

    for p in P:
        solver.Add(ST[p] +Dur[p] <= FT[p])
        
    for p in P:
        solver.Add(ST[p] >= SD[p]) 
    
    for p in P:
        solver.Add(FT[p] <= FD[p]) 

    for w in W:
        solver.Add(sum(X[p, w]* Dur[p] for p in P) <= WAD[w])
        
    # Add the constraint to the solver
    for p1 in P:
        for p2 in P:
            # Safely access the Prec dictionary
            precedence_value = Prec.get((p2, p1), 0)  # Default to 0 if no precedence value is defined
            solver.Add(FT[p1] * precedence_value <= ST[p2])


                                                    
    weight_ft  = 0.5         
    weight_score = 0.3           
    weight_cal = 0.5     
    weight_cost = 0.1    

    # Objective 1: Minimize Sum(FT[p] for p in P)
    objective1 = sum(FT[p] for p in P)

    # Objective 2: Minimize Sum(X[p, w] * Score[w] for p in P for w in W)
    objective2 = sum(X[p, w] * Score[w] for p in P for w in W)

    # Objective 3: Maximize Sum(X[p, w] * WAD[w] for p in P for w in W) (Maximization converted to minimization by negating)
    objective3 = sum(X[p, w] *(1- Cal[w,t]) for p in P for w in W for t in T)

    # Objective 4: Minimize Sum(X[p, w] * Cost[w] for p in P for w in W)
    objective4 = sum(X[p, w] * Cost_sheet[w] for p in P for w in W)

    # Define the combined objective with relative weights
    combined_objective = (weight_ft  * objective1 +
                        weight_score * objective2 +
                        weight_cal * objective3 +
                        weight_cost * objective4)

    # Minimize the combined objective
    solver.Minimize(combined_objective)
    status = solver.Solve()

    # Check the status and extract the results
    if status == pywraplp.Solver.OPTIMAL:
        print('Solution found!')
        # Extract the solution values
        for p in P:
            for w in W:
                    if X[p, w].solution_value() == 1:    
                        print(f'Project {p+1}, Worker {w+1}')
        for p in P:
            print(f'Project {p+1}, ST: {ST[p].solution_value()}, FT: {FT[p].solution_value()}, Due date: {SD[p]}, Dur: {Dur[p]}')
            
        print('Objective value =', solver.Objective().Value())
    else:
        print('No optimal solution found.')

    print("Start Heuristics") 

    # Initialize dictionaries for adjusted start and finish times
    adjusted_ST = {}
    adjusted_FT = {}

    # Initialize a dictionary to track the latest finish time for each worker
    latest_finish_time = {w: 0 for w in W}  # For each worker, store their latest finish time
    for p in P:
        # Find the worker assigned to project p
        for w in W:
            # Ensure we use the value of x[p, w] correctly (as a Python boolean check)
            if X[p, w].solution_value() == 1:  # Get the assigned worker from the solver's solution value
                assigned_worker = int(w) # Get the assigned worker for this project
                break  # Stop once the worker is found

        # Get the required start time and duration from the project variables
        required_start_time = ST[p].solution_value()  # Extract solution values for start time
        duration = Dur[p]  # Duration of the project is a parameter, so it's directly accessible
        available_days = 0  # Counter for available workdays

        # Assign the earliest possible start time based on the finish time of the last project
        earliest_possible_start = max(latest_finish_time[assigned_worker], required_start_time)

        # Start checking the worker's calendar from the earliest possible start time
        t = earliest_possible_start

        # Find the first available day on or after the earliest possible start time
        while Cal[assigned_worker, t] == 0:  # If it's an off day, move to the next day
            t += 1

        # Set the adjusted start time to the first available day
        adjusted_ST[p] = t

        # Now, calculate the finish time based on worker availability and project duration
        while available_days < duration:  # Work for the required number of available days
            if Cal[assigned_worker, t] == 1:  # If the worker is available on day t
                available_days += 1  # Increment available workdays
            t += 1  # Move to the next time period

        # The finish time is the last day the worker is available for the full duration
        adjusted_FT[p] = t - 1  # t-1 because t was incremented after the last workday

        # Update the latest finish time for the assigned worker so the next project starts after this
        latest_finish_time[assigned_worker] = adjusted_FT[p]

    # Output adjusted start and finish times
    for p in P:
        print(f"Project {p+1}: Adjusted ST: {adjusted_ST[p]}, Adjusted FT: {adjusted_FT[p]} ")
    
    
    import datetime

    # Define current date as the base for ST[1]
    current_date = datetime.datetime.now().date()

    # Assuming adjusted_ST and adjusted_FT contain the start and finish times as integers (days offset from ST[1])
    # Convert those to dates
    adjusted_ST_dates = {p: current_date + datetime.timedelta(days=adjusted_ST[p]) for p in P}
    adjusted_FT_dates = {p: current_date + datetime.timedelta(days=adjusted_FT[p]) for p in P}

    # Create the DataFrame with the updated dates
    data = {
        "Project": [f"Project {p+1}" for p in P],
        # Extract the worker assigned to each project by checking the solution value of X[p, w]
        "Worker Assigned": [
            next((w +1 for w in W if X[p, w].solution_value() == 1), "No Worker Assigned") for p in P
        ],
        "Adjusted ST": [adjusted_ST_dates.get(p, "No ST") for p in P],  # Start date as date
        "Adjusted FT": [adjusted_FT_dates.get(p, "No FT") for p in P] ,  # Finish date as date
        "Project Duration": [Dur[p] for p in P]
    }

    df = pd.DataFrame(data)

    def get_skills_for_project(p):
        skills = [f"Skill {s+1}" for s in range(len(unique_skills)) if RSkl.get((s, p), 0) == 1]
        return ", ".join(skills) if skills else "No Skills"

    df["Skills Required"] = [get_skills_for_project(p) for p in P]



    # Save to Excel
    file_path = 'project_schedule_with_dates.xlsx'
    df.to_excel(file_path, index=False)

    print(f"Excel file saved at: {file_path}")

    # Graphs
                                    # WAD vs Available Time
    Sum_X_Dur_list = []
    for w in W:
            Sum_X_Dur = 0
            for p in P:
                # Retrieve the value of the decision variable X[p, w] from the solution
                if X[(p, w)].solution_value() == 1:
                    Sum_X_Dur += X[(p, w)].solution_value() * Dur[p]
                    # print(f'X[{p},{w}] * Dur[{p}] = {X[(p, w)].solution_value()} * {Dur[p]} = {X[(p, w)].solution_value() * Dur[p]}')
            Sum_X_Dur_list.append(Sum_X_Dur)
    print("Total duration assigned to each worker:", Sum_X_Dur_list)
                                
    WAD_list = [WAD[w] for w in W if w in WAD]
    print(WAD_list)

    workers = np.arange(len(WAD_list))

    # Set bar width
    bar_width = 0.35

    # Create a figure and axis with increased size
    fig, ax = plt.subplots(figsize=(10, 6))  # Increased figure size

    # Plot the available time (WAD) as the first bar for each worker
    bar1 = ax.bar(workers, WAD_list, bar_width, label='Available Time (WAD)')

    # Plot the utilized time (Sum_X_Dur) as the second bar for each worker, slightly offset to the right
    bar2 = ax.bar(workers + bar_width, Sum_X_Dur_list, bar_width, label='Utilized Time')

    # Set labels and title
    ax.set_xlabel('Workers')
    ax.set_ylabel('Time (in days)')
    ax.set_title('Comparison of Available vs Utilized Time for Each Worker')

    # Set x-axis ticks at the center of the bars and label them as workers
    ax.set_xticks(workers + bar_width / 2)
    ax.set_xticklabels([f'W {i+1}' for i in workers], rotation=45, ha="right")

    # Adjust the layout to add more space between elements
    plt.subplots_adjust(bottom=0.2, left=0.1, right=0.9, top=0.9)

    # Add legend
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=2)
    
    # Add values on top of each bar (for available time)
    for i, v in enumerate(WAD_list):
        ax.text(i - 0.05, v + 0.5, str(v), color='blue', fontweight='bold')

    # Add values on top of each bar (for utilized time)
    for i, v in enumerate(Sum_X_Dur_list):
        ax.text(i + bar_width - 0.05, v + 0.5, str(v), color='orange', fontweight='bold')
        
    # Display the plot
    plt.tight_layout()
    print("showing")
    plt.show()

                                # Cost per Worker

    Cost_W = []
    for w in W:
            Cost_Project = 0
            for p in P:
                if X[(p, w)].solution_value() == 1:
                    Cost_Project += X[(p, w)].solution_value() *Cost[w]*Dur[p]
            Cost_W.append(Cost_Project)
    print("Total Cost to each worker:", Cost_W)

    Cost_P = []
    for p in P:
            Cost_Project = 0
            for w in W:
                if X[(p, w)].solution_value() == 1:
                    Cost_Project += X[(p, w)].solution_value() *Cost[w]*Dur[p]
            Cost_P.append(Cost_Project)
    print("Total Cost to each project:", Cost_P)


    fig, ax = plt.subplots(figsize=(10, 6))

    # Create the bar graph
    ax.bar(W, Cost_W, color='#76C7C0', edgecolor='blue')

    # Set labels and title
    ax.set_xlabel('Workers')
    ax.set_ylabel('Total Cost')
    ax.set_title('Total Cost per Worker')

    # Set x-axis ticks and labels for workers
    ax.set_xticks(np.arange(len(W)))
    ax.set_xticklabels([f'Worker {i+1}' for i in W])

    # Add total cost labels on top of each bar
    for i, total_cost in enumerate(Cost_W):
        ax.text(i, total_cost + 2, f'{total_cost}', ha='center', fontweight='bold')

    # Display the plot
    plt.tight_layout()
    plt.show()


                                    # Cost per Project
                                    
    fig, ax = plt.subplots(figsize=(12, 6))   

    # Create the bar graph
    ax.bar(P, Cost_P, color='#FFBB28', edgecolor='blue')

    # Set labels and title
    ax.set_xlabel('Projects')
    ax.set_ylabel('Cost')
    ax.set_title('Cost per Project')

    # Rotate x-axis labels for better spacing
    ax.set_xticks(np.arange(len(P)))
    ax.set_xticklabels([f'Project {i+1}' for i in P], rotation=45, ha="right")

    # Add cost labels on top of each bar
    for i, cost in enumerate(Cost_P):
        ax.text(i, cost + 1, f'{cost}', ha='center', fontweight='bold')

    # Add padding to prevent overlap
    plt.tight_layout()

    # Display the plot
    plt.show()
                                    # Worker Assignment

    for p in P:
        for w in W:
            if X[p, w].solution_value() == 1:  
                print(f"X[{p+1}, {w+1}] = 1")

    x_values = []
    y_values = []            
    for p in P:
        for w in W:
            if X[p, w].solution_value() == 1:  # Only include assignments where X[p, w] is 1
                x_values.append(w)  # Worker on x-axis
                y_values.append(p)  # Project on y-axis

    # Create the scatter plot
    plt.figure(figsize=(10, 6))
    plt.scatter(x_values, y_values, color='blue', marker='o')

    # Set labels and title
    plt.xlabel('Workers')
    plt.ylabel('Projects')
    plt.title('Project Assignment to Worker')

    # Set x-axis and y-axis ticks
    plt.xticks(W, [f'Worker {w+1}' for w in W])
    plt.yticks(P, [f'Project {p+1}' for p in P])

    # Show the plot            
    plt.tight_layout()
    plt.show()
                                            # Gantt Chart

    adjusted_ST_dates = {p: current_date + datetime.timedelta(days=adjusted_ST[p]) for p in P}
    adjusted_FT_dates = {p: current_date + datetime.timedelta(days=adjusted_FT[p]) for p in P}

    gantt_data = []
    for p in P:
        gantt_data.append({
            'Project': p,
            'Start': adjusted_ST_dates[p],
            'Finish': adjusted_FT_dates[p]
        })

    # Convert to DataFrame for plotting
    df = pd.DataFrame(gantt_data)

    # Create Gantt chart
    fig, ax = plt.subplots(figsize=(10, 5))

    # Plot each project as a horizontal bar
    for index, row in df.iterrows():
        ax.barh(row['Project'], (row['Finish'] - row['Start']).days, left=row['Start'], align='center')

    # Format the x-axis as dates
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=5))
    plt.xticks(rotation=45)

    # Set y-axis as sequential numbers
    ax.set_yticks(range(1, len(P)+1))  # Sequential numbering from 1 to the number of projects
    ax.set_ylabel('Project No')

    # Labels and title
    ax.set_xlabel('Time')
    ax.set_title('Optimized Gantt Chart')

    plt.tight_layout()
    plt.show()

# optimize(7)

@login_required
def projects(request):
    
    
    projects = Project.objects.filter(user=request.user).order_by('id')
    return render(request, 'dashboard/projects.html', {'projects':projects})


@login_required
def constraints(request, project_id):
    project = Project.objects.get(id=project_id)

    if request.method == 'POST':
        # Extract Sets Data
        sets_data = []
        projects = request.POST.getlist('projects')
        teamleads = request.POST.getlist('teamleads')
        timeperiod = request.POST.getlist('timeperiod')
        skills = request.POST.getlist('skills')
        managers = request.POST.getlist('managers')

        # Ensure all lists have the same length
        min_length = min(len(projects), len(teamleads), len(timeperiod), len(skills), len(managers))
        
        for i in range(min_length):
            sets_data.append(Set(
                project=project,
                projects=projects[i],
                teamleads=teamleads[i],
                timeperiod=timeperiod[i],
                skills=skills[i],
                managers=managers[i]
            ))

        # Save Sets Data to Database
        Set.objects.bulk_create(sets_data)

        # Extract Skills Data
        skills_data = []
        skill_values = request.POST.getlist('skills')
        teamleads_values = request.POST.getlist('teamleads')
        value_values = request.POST.getlist('value')

        # Ensure all lists have the same length
        min_length = min(len(skill_values), len(teamleads_values), len(value_values))
        
        for i in range(min_length):
            skills_data.append(Skill(
                project=project,
                skills=skill_values[i],
                teamleads=teamleads_values[i],
                value=value_values[i]
            ))

        # Save Skills Data to Database
        Skill.objects.bulk_create(skills_data)

        # Extract Duration Data
        duration_data = []
        project_values = request.POST.getlist('projects')
        value_values = request.POST.getlist('value')

        # Ensure all lists have the same length
        min_length = min(len(project_values), len(value_values))
        
        for i in range(min_length):
            duration_data.append(Duration(
                project=project,
                projects=project_values[i],
                value=value_values[i]
            ))

        # Save Duration Data to Database
        Duration.objects.bulk_create(duration_data)

        # Create Excel file
        # Fetch data from the database for the specific project
        set_data = Set.objects.filter(project=project)
        skill_data = Skill.objects.filter(project=project)
        duration_data = Duration.objects.filter(project=project)

        # Convert querysets to DataFrames
        set_df = pd.DataFrame.from_records(set_data.values())
        skill_df = pd.DataFrame.from_records(skill_data.values())
        duration_df = pd.DataFrame.from_records(duration_data.values())

        # Prepare file name
        file_name = f'{project.name}_constraints.xlsx'
        file_path = os.path.join('dashboard/files', file_name)  # Replace 'path/to/save/directory' with the desired path

        # Write data to Excel with multiple sheets
        with pd.ExcelWriter(file_path) as writer:
            set_df.to_excel(writer, sheet_name='Sets', index=False)
            skill_df.to_excel(writer, sheet_name='Skills', index=False)
            duration_df.to_excel(writer, sheet_name='Duration', index=False)

        # Optionally, send file as response to download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            set_df.to_excel(writer, sheet_name='Sets', index=False)
            skill_df.to_excel(writer, sheet_name='Skills', index=False)
            duration_df.to_excel(writer, sheet_name='Duration', index=False)

        return response

    return render(request, 'dashboard/constrains_input.html', {'project': project})






from django.http import HttpResponseForbidden
import logging
import datetime
logger = logging.getLogger(__name__)
@login_required
def process_data(request):
    if request.method == 'POST':
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            project_id = request.POST.get('project_id')

            # Fetch the project, or return a 404 if not found
            current_project = get_object_or_404(Project, id=project_id)
            existing_files = ProjectFiles.objects.filter(project=current_project)
            for file in existing_files:
                file.file.delete()  # Delete the physical file
                file.delete() 
            # Check if the user has permission to modify this project
            if current_project.user != request.user:
                return HttpResponseForbidden("You are not allowed to modify this project.")

            try:
                # Read the uploaded Excel file
                all_sheets = pd.read_excel(excel_file, sheet_name=None, header=None)
            except Exception as e:
                logger.error(f"Error processing file for project {project_id}: {e}")
                return render(request, 'dashboard/upload.html', {
                    'error': f"Error processing file: {e}"
                })

            sheet_names = list(all_sheets.keys())

            Sets    = all_sheets[sheet_names[0]] 
            W_skills_Avl   = all_sheets[sheet_names[1]] 
            Duration     = all_sheets[sheet_names[2]]
            Req_skills   = all_sheets[sheet_names[3]]
            Manager_skills_Avl    = all_sheets[sheet_names[4]]
            Due_Starting_Date     = all_sheets[sheet_names[5]]
            Due_Finishing_Date   = all_sheets[sheet_names[6]]
            # WAD_sheet            = all_sheets[sheet_names[7]]
            # Score_sheet          = all_sheets[sheet_names[8]]
            Cost_sheet           = all_sheets[sheet_names[9]]    
            Prec_sheet = all_sheets[sheet_names[10]]
            calendar_df = all_sheets[sheet_names[11]]

            # print(Calendar)

            projects = Sets.iloc[1, 0]   # Second row, first column (Projects)
            teamleads = Sets.iloc[1, 1]  # Second row, second column (Teamleads)
            timeperiod = Sets.iloc[1, 2]  # Second row, third column (Timeperiod)
            skills = Sets.iloc[1, 3]    # Second row, fourth column (Skills)
            start_date = Sets.iloc[1, 4]  # Second row, fifth column (Start Period)
            finish_date = Sets.iloc[1, 5]  # Second row, sixth column (Finish Period)

            # Display the extracted values
            print( projects)
            print( teamleads)
            print( timeperiod)
            print( skills)
            print( start_date)
            print( finish_date)


            # Sets
            T = range(timeperiod)  # time periods, e.g., range(24) for 24 hours or days
            S = range(skills)        # skills
            W = range(teamleads)     # workforce    
            P = range(projects)      # Projects  
            
            # sys.exit()

            
            calendar_df.reset_index(drop=True, inplace=True)


            Cal = {}

            for i in range(calendar_df.shape[0]):  
                for j in range(calendar_df.shape[1]): 
                    value = calendar_df.iloc[i, j]  
                    
                    # Adjust index manually
                    adjusted_i = i - 1
                    adjusted_j = j - 1

                    if isinstance(value, (int, float)) and not pd.isna(value):
                        Cal[(adjusted_i, adjusted_j)] = int(value)

            print('Calender',Cal)

            Prec_sheet.reset_index(drop=True, inplace=True)

            Prec = {}

            for i in range(Prec_sheet.shape[0]):  
                for j in range(Prec_sheet.shape[1]): 
                    value = Prec_sheet.iloc[i, j]  
                    
                    # Adjust index manually
                    adjusted_i = i - 1
                    adjusted_j = j - 1

                    if isinstance(value, (int, float)) and not pd.isna(value):
                        Prec[(adjusted_i, adjusted_j)] = int(value)

            print('Precedence Matrix',Prec)



            skills_W = W_skills_Avl.iloc[1:, 0]   
            workers = W_skills_Avl.iloc[1:, 1]  
            values_W = W_skills_Avl.iloc[1:, 2]  
            
            unique_skills  = skills_W.unique()
            unique_workers = workers.unique()

            WSAvl  = {}

            # Populate the dictionary with values from the DataFrame
            for i in range(len(skills_W )):
                skill = skills_W.iloc[i]
                worker = workers.iloc[i]
                value = values_W.iloc[i]
                
                # Find the index of the current skill and project
                skill_index = list(unique_skills).index(skill)
                Worker_index = list(unique_workers).index(worker)
                
                # Add the value to the dictionary
                WSAvl[(skill_index , Worker_index )] = int(value)

            # print(WSAvl )

            # sys.exit()

            skills_Req = Req_skills.iloc[1:, 0]   # Skills column
            projects_Req = Req_skills.iloc[1:, 1]  # Projects column
            values_Req = Req_skills.iloc[1:, 2]    # Values column

            # Get unique skills and projects
            unique_skills_req = skills_Req.unique()
            unique_projects_req = projects_Req.unique()

            # Create the dictionary for storing requirement values
            RSkl = {}

            # Populate the dictionary with values from the DataFrame
            for i in range(len(skills_Req)):
                skill = skills_Req.iloc[i]
                project = projects_Req.iloc[i]
                value = values_Req.iloc[i]
                
                # Find the index of the current skill and project
                skill_index = list(unique_skills_req).index(skill)
                project_index = list(unique_projects_req).index(project)
                
                # Add the value to the dictionary
                RSkl[(skill_index ,project_index )] = int(value)

            # print(RSkl)

            skills_M = Manager_skills_Avl.iloc[1:, 0]   # Skills column
            Manager = Manager_skills_Avl.iloc[1:, 1]    # Managers column
            values = Manager_skills_Avl.iloc[1:, 2]     # Values column

            # Get unique managers and skills
            unique_skills = skills_M.unique()
            unique_managers = Manager.unique()

            # Create the dictionary for storing availability values
            MSAvl  = {}

            # Populate the dictionary with values from the DataFrame
            for i in range(len(skills_M)):
                skill = skills_M.iloc[i]
                manager = Manager.iloc[i]
                value = values.iloc[i]
                
                # Find the index of the current skill and manager
                skill_index = list(unique_skills).index(skill)
                manager_index = list(unique_managers).index(manager)
                
                # Add the value to the dictionary
                MSAvl[(skill_index, manager_index )] = int(value)


            projects_dur = Duration.iloc[1:, 0]   
            values_dur = Duration.iloc[1:, 1]  
            Dur = {i: values_dur.iloc[i] for i in range(len(projects_dur))}

            
            projects_sd = Due_Starting_Date.iloc[1:, 0]  # Project names from Due_Starting_Date
            values_sd = Due_Starting_Date.iloc[1:, 1]  # Start dates from Due_Starting_Date
            SD = {i: values_sd.iloc[i] for i in range(len(projects_sd))}



            projects_fd = Due_Finishing_Date.iloc[1:, 0]  # Project names from Due_Starting_Date
            values_fd = Due_Finishing_Date.iloc[1:, 1]  # Start dates from Due_Starting_Date
            FD = {i: values_fd.iloc[i] for i in range(len(projects_fd))}

            

            
            # WAD= WAD_sheet .iloc[1:, 0]  # Project names from Due_Starting_Date
            # values_WAD = WAD_sheet.iloc[1:, 1]  # Start dates from Due_Starting_Date
            # WAD = {i: values_WAD.iloc[i] for i in range(len(WAD))} 
            # print("WAD values", WAD)

            
            
            calendar_df_dropped = calendar_df.iloc[1:, 1:]
            WAD_values = calendar_df_dropped.iloc[:, T].sum(axis=1).to_list() 
            WAD = {i: WAD_values[i] for i in range(len(WAD_values))}
            print("WAD values  :", WAD)
            

            
            Score = {}
            for w in W:
                score_sum = 0
                for t in T:
                    score_sum += (len(T) - t + 1 - Cal[w, t]) * (1 - Cal[w, t])
                Score[w] = score_sum / len(T)
            print("Workers Score", Score)


            Cost=  Cost_sheet .iloc[1:, 0]  # Project names from Due_Starting_Date
            values_Cost = Cost_sheet.iloc[1:, 1]  # Start dates from Due_Starting_Date
            Cost = {i: values_Cost.iloc[i] for i in range(len(Cost))} 
            
            # print("Workers Available Days",WAD)
            # print("Workers Availablity Score",Score)
            # print("Workers Cost per day",Cost)


            # sys.exit()

            print("Start Solving")

            X =  {(p, w): solver.BoolVar(f'X[{p}][{w}]') for p in P for w in W}
            ST = {(p): solver.IntVar(0, 500, f'ST[{p}]') for p in P }
            FT = {(p): solver.IntVar(0, 500, f'FT[{p}]') for p in P }

            # Constraints

            for p in P:
                solver.Add(sum(X[p, w] for w in W) == 1)
                
            for p in P:
                for w in W:
                    for s in S:
                        solver.Add( RSkl[s,p] - WSAvl[s, w] <= 2*(1-X[p, w]))
                        
            for p in P:
                solver.Add(ST[p] +Dur[p] <= FT[p])
                
            for p in P:
                solver.Add(ST[p] >= SD[p]) 
            
            for p in P:
                solver.Add(FT[p] <= FD[p])  
                
            max_time = max(T)
            # for p in P:
            #     solver.Add(FT[p] <= max_time)
            

            for w in W:
                solver.Add(sum(X[p, w]* Dur[p] for p in P) <= WAD[w])
                
            for p1 in P:
                for p2 in P:
                    solver.Add(FT[p1]*Prec[p2,p1] <= ST[p2])

                                                            
            weight_ft  = 1 
            weight_score = 1          
            weight_cal = 1     
            weight_cost = 1    

            # Objective 1: Minimize Sum(FT[p] for p in P)
            objective1 = sum(FT[p] for p in P)

            # # Objective 2: Minimize Sum(X[p, w] * Score[w] for p in P for w in W)
            objective2 = sum(X[p, w] * Score[w] for p in P for w in W)

            # # Objective 3: Maximize Sum(X[p, w] * WAD[w] for p in P for w in W) (Maximization converted to minimization by negating)
            # objective3 = sum(X[p, w] *(1- Cal[w,t]) for p in P for w in W for t in T)

            # Objective 4: Minimize Sum(X[p, w] * Cost[w] for p in P for w in W)
            objective4 = sum(X[p, w] * Cost[w] for p in P for w in W)

            # Define the combined objective with relative weights
            combined_objective = (weight_ft  * objective1 +
                                weight_score * objective2 +
                                # weight_cal * objective3 +
                                weight_cost * objective4 
                                )

            # Minimize the combined objective
            solver.Minimize(combined_objective)
            status = solver.Solve()

            # Check the status and extract the results
            if status == pywraplp.Solver.OPTIMAL:
                print('\nSolution found!')
                # Extract the solution values
                for p in P:
                    for w in W:
                            if X[p, w].solution_value() == 1:    
                                print(f'Project {p+1}, Worker {w+1}')
                for p in P:
                    print(f'Project {p+1}, ST: {ST[p].solution_value()}, FT: {FT[p].solution_value()}, Due date: {SD[p]}, Dur: {Dur[p]}')
                    

                
                print('\nObjective value =', solver.Objective().Value())

                objective1_value = sum(FT[p].solution_value() for p in P)
                print(f'Objective 1 (Sum of FT): {objective1_value}')
                
                # Objective 2: Minimize Sum(X[p, w] * Score[w] for p in P for w in W)
                objective2_value = sum(X[p, w].solution_value() * Score[w] for p in P for w in W)
                print(f'Objective 2 (Sum of X[p, w] * Score[w]): {objective2_value}')
                
                # # Objective 3: Maximize Sum(X[p, w] * WAD[w] for p in P for w in W) (Maximization converted to minimization by negating)
                # objective3_value = sum(X[p, w].solution_value() * (1 - Cal[w, t]) for p in P for w in W for t in T)
                # print(f'Objective 3 (Maximized Availability): {objective3_value}')
                
                # Objective 4: Minimize Sum(X[p, w] * Cost[w] for p in P for w in W)
                objective4_value = sum(X[p, w].solution_value() * Cost[w] for p in P for w in W)
                print(f'Objective 4 (Sum of X[p, w] * Cost[w]): {objective4_value}')

            else:
                print('No optimal solution found.')

            print("\nStart Heuristics") 

            # Initialize dictionaries for adjusted start and finish times
            adjusted_ST = {}
            adjusted_FT = {}
            
            # Initialize a dictionary to track the latest finish time for each worker
            latest_finish_time = {w: 0 for w in W}  # For each worker, store their latest finish time

            for p in P:
                # Find the worker assigned to project p
                for w in W:
                    # Ensure we use the value of x[p, w] correctly (as a Python boolean check)
                    if X[p, w].solution_value() == 1:  # Get the assigned worker from the solver's solution value
                        assigned_worker = int(w)  # Get the assigned worker for this project
                        break  # Stop once the worker is found

                # Get the required start time and duration from the project variables
                required_start_time = ST[p].solution_value()  # Extract solution values for start time
                duration = Dur[p]  # Duration of the project is a parameter, so it's directly accessible
                available_days = 0  # Counter for available workdays

                # Assign the earliest possible start time based on the finish time of the last project
                earliest_possible_start = max(latest_finish_time[assigned_worker], required_start_time)

                # Start checking the worker's calendar from the earliest possible start time
                t = int(earliest_possible_start)

                # Find the first available day on or after the earliest possible start time
                while Cal[assigned_worker, t] == 0:  # If it's an off day, move to the next day
                    t += 1

                # Set the adjusted start time to the first available day
                adjusted_ST[p] = t

                # Now, calculate the finish time based on worker availability and project duration
                while available_days < duration:  # Work for the required number of available days
                    if Cal[assigned_worker, t] == 1:  # If the worker is available on day t
                        available_days += 1  # Increment available workdays
                    t += 1  # Move to the next time period

                # The finish time is the last day the worker is available for the full duration
                adjusted_FT[p] = t - 1  # t-1 because t was incremented after the last workday

                # Update the latest finish time for the assigned worker so the next project starts after this
                latest_finish_time[assigned_worker] = adjusted_FT[p]
                T =50
                if adjusted_FT[p] > T:
                    print(f"Project {p} exceeds time period {T}: Start Time = {adjusted_ST[p]}, Finish Time = {adjusted_FT[p]}")


            # sys.exit()
            # Define current date as the base for ST[1]
            current_date = datetime.datetime.now().date()

            # Assuming adjusted_ST and adjusted_FT contain the start and finish times as integers (days offset from ST[1])
            # Convert those to dates
            adjusted_ST_dates = {p: current_date + datetime.timedelta(days=adjusted_ST[p]) for p in P}
            adjusted_FT_dates = {p: current_date + datetime.timedelta(days=adjusted_FT[p]) for p in P}

            # Create the DataFrame with the updated dates
            data = {
                    
                "Project": [f"Project {p+1}" for p in P],
                "Start Time (ST)": [ST[p].solution_value() for p in P],
                "Finish Time (FT)": [FT[p].solution_value() for p in P],
                "Due Date (SD)": [SD[p] for p in P],
                "Duration (Dur)": [Dur[p] for p in P],
                "Adjusted Start Time": [adjusted_ST[p] for p in P],
                "Adjusted Finish Time": [adjusted_FT[p] for p in P],
                
            
                # Extract the worker assigned to each project by checking the solution value of X[p, w]
                "Worker Assigned": [
                    next((w +1 for w in W if X[p, w].solution_value() == 1), "No Worker Assigned") for p in P
                ],
                "ST Date": [adjusted_ST_dates.get(p, "No ST") for p in P],  # Start date as date
                "FT Date": [adjusted_FT_dates.get(p, "No FT") for p in P]   # Finish date as date
            
            }

            df = pd.DataFrame(data)

            def get_skills_for_project(p):
                skills = [f"Skill {s+1}" for s in range(len(unique_skills)) if RSkl.get((s, p), 0) == 1]
                return ", ".join(skills) if skills else "No Skills"

            df["Skills Required"] = [get_skills_for_project(p) for p in P]



            # Save to Excel
            file_path = 'combined_project_scheduling.xlsx'
            df.to_excel(file_path, index=False)

            print(f"Excel file saved at: {file_path}")


            input_file = excel_file
            cal_df = pd.read_excel(input_file, sheet_name='Calender', index_col=0)

            # Initialize a copy of the calendar to store assignments
            output_df = cal_df.copy()

            for p in P:
                for w in W:
                    if X[p, w].solution_value() == 1:  # Worker is assigned to project p
                        start_day = int(adjusted_ST[p])
                        finish_day = int(adjusted_FT[p])
                        # Replace 1's in the worker's calendar with the project name for the assigned days
                        for day in range(start_day, finish_day + 1):
                            day_col = output_df.columns[day - 1]  # Get the day column name (Excel's date columns)
                            if cal_df.at[f'TL{w+1}', day_col] == 1:  # Check availability in the calendar
                                output_df.at[f'TL{w+1}', day_col] = f"Project {p+1}"

            # Replace remaining 1's with "Not assigned" and 0's with "Unavailable"
            output_df = output_df.replace(1, "NA")
            output_df = output_df.replace(0, "UA")

            # Assuming `output_df` is the DataFrame you want to write to Sheet2
            output_file = 'combined_project_scheduling.xlsx'

            # Load the existing workbook to preserve Sheet1
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
                # Write the new DataFrame to a new sheet ('Sheet2') without altering 'Sheet1'
                output_df.to_excel(writer, sheet_name='Sheet2')

            print(f"Data has been successfully written to 'Sheet2' in {output_file}")
            


            # import pandas as pd
            # from datetime import datetime, timedelta  # Correct import of timedelta

            # # Load the calendar DataFrame
            # input_file = 'Input.xlsx'
            # cal_df = pd.read_excel(input_file, sheet_name='Calender', index_col=0)

            # # Initialize a copy of the calendar to store assignments
            # output_df = cal_df.copy()

            # # Generate date labels for the columns, starting from the current date
            # current_date = datetime.now().date()  # This gets the current date (no time component)
            # date_labels = [current_date + timedelta(days=i) for i in range(output_df.shape[1])]  # List of dates

            # # Convert date_labels to string format (if your original `cal_df` columns are strings)
            # date_labels_str = [str(date) for date in date_labels]

            # # Ensure that both the original `cal_df` and `output_df` have aligned date columns
            # output_df.columns = date_labels_str
            # cal_df.columns = date_labels_str

            # # Assign workers to projects based on the solution
            # for p in P:
            #     for w in W:
            #         if X[p, w].solution_value() == 1:  # Worker is assigned to project p
            #             start_day = int(adjusted_ST[p])  # Adjusted start day (int)
            #             finish_day = int(adjusted_FT[p])  # Adjusted finish day (int)
            #             # Replace 1's in the worker's calendar with the project name for the assigned days
            #             for day in range(start_day, finish_day + 1):
            #                 day_col = output_df.columns[day - 1]  # Get the corresponding day column
            #                 if cal_df.at[f'TL{w+1}', day_col] == 1:  # Check availability in the calendar
            #                     output_df.at[f'TL{w+1}', day_col] = f"Project {p+1}"

            # # Replace remaining 1's with "NA" (Not Assigned) and 0's with "UA" (Unavailable)
            # output_df = output_df.replace(1, "NA")
            # output_df = output_df.replace(0, "UA")

            # # Save the updated calendar DataFrame to a new Excel file
            # output_file = 'updated_worker_calendar.xlsx'
            # output_df.to_excel(output_file)

            # print(f"Data has been successfully saved to {output_file}")



            # Graphs
                                            # WAD vs Available Time

            print("\nGraphs")                                
            Sum_X_Dur_list = []
            for w in W:
                    Sum_X_Dur = 0
                    for p in P:
                        # Retrieve the value of the decision variable X[p, w] from the solution
                        if X[(p, w)].solution_value() == 1:
                            Sum_X_Dur += X[(p, w)].solution_value() * Dur[p]
                            # print(f'X[{p},{w}] * Dur[{p}] = {X[(p, w)].solution_value()} * {Dur[p]} = {X[(p, w)].solution_value() * Dur[p]}')
                    Sum_X_Dur_list.append(Sum_X_Dur)
            print("Total duration assigned to each worker:", Sum_X_Dur_list)
                                        
            WAD_list = [WAD[w] for w in W if w in WAD]
            print(WAD_list)

            workers = np.arange(len(WAD_list))

            # Set bar width
            bar_width = 0.35

            # Create a figure and axis with increased size
            fig, ax = plt.subplots(figsize=(10, 6))  # Increased figure size

            # Plot the available time (WAD) as the first bar for each worker
            bar1 = ax.bar(workers, WAD_list, bar_width, label='Available Time (WAD)')

            # Plot the utilized time (Sum_X_Dur) as the second bar for each worker, slightly offset to the right
            bar2 = ax.bar(workers + bar_width, Sum_X_Dur_list, bar_width, label='Utilized Time')

            # Set labels and title
            ax.set_xlabel('Workers')
            ax.set_ylabel('Time (in days)')
            ax.set_title('Comparison of Available vs Utilized Time for Each Worker')

            # Set x-axis ticks at the center of the bars and label them as workers
            ax.set_xticks(workers + bar_width / 2)
            ax.set_xticklabels([f'W {i+1}' for i in workers], rotation=45, ha="right")

            # Adjust the layout to add more space between elements
            plt.subplots_adjust(bottom=0.2, left=0.1, right=0.9, top=0.9)

            # Add legend
            ax.legend(loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=2)
            
            # Add values on top of each bar (for available time)
            for i, v in enumerate(WAD_list):
                ax.text(i - 0.05, v + 0.5, str(v), color='blue', fontweight='bold')

            # Add values on top of each bar (for utilized time)
            for i, v in enumerate(Sum_X_Dur_list):
                ax.text(i + bar_width - 0.05, v + 0.5, str(v), color='orange', fontweight='bold')
                
            # Display the plot
            plt.tight_layout()
            plt.show()

                                        # Cost per Worker

            Cost_W = []
            for w in W:
                    Cost_Project = 0
                    for p in P:
                        if X[(p, w)].solution_value() == 1:
                            Cost_Project += X[(p, w)].solution_value() *Cost[w]*Dur[p]
                    Cost_W.append(Cost_Project)
            print("Total Cost to each worker:", Cost_W)

            Cost_P = []
            for p in P:
                    Cost_Project = 0
                    for w in W:
                        if X[(p, w)].solution_value() == 1:
                            Cost_Project += X[(p, w)].solution_value() *Cost[w]*Dur[p]
                    Cost_P.append(Cost_Project)
            print("Total Cost to each project:", Cost_P)


            fig, ax = plt.subplots(figsize=(10, 6))

            # Create the bar graph
            ax.bar(W, Cost_W, color='#76C7C0', edgecolor='blue')

            # Set labels and title
            ax.set_xlabel('Workers')
            ax.set_ylabel('Total Cost')
            ax.set_title('Total Cost per Worker')

            # Set x-axis ticks and labels for workers
            ax.set_xticks(np.arange(len(W)))
            ax.set_xticklabels([f'Worker {i+1}' for i in W])

            # Add total cost labels on top of each bar
            for i, total_cost in enumerate(Cost_W):
                ax.text(i, total_cost + 2, f'{total_cost}', ha='center', fontweight='bold')

            # Display the plot
            plt.tight_layout()
            plt.show()


                                            # Cost per Project
                                            
            fig, ax = plt.subplots(figsize=(12, 6))   

            # Create the bar graph
            ax.bar(P, Cost_P, color='#FFBB28', edgecolor='blue')

            # Set labels and title
            ax.set_xlabel('Projects')
            ax.set_ylabel('Cost')
            ax.set_title('Cost per Project')

            # Rotate x-axis labels for better spacing
            ax.set_xticks(np.arange(len(P)))
            ax.set_xticklabels([f'Project {i+1}' for i in P], rotation=45, ha="right")

            # Add cost labels on top of each bar
            for i, cost in enumerate(Cost_P):
                ax.text(i, cost + 1, f'{cost}', ha='center', fontweight='bold')

            # Add padding to prevent overlap
            plt.tight_layout()

            # Display the plot
            plt.show()
                                            # Worker Assignment

            for p in P:
                for w in W:
                    if X[p, w].solution_value() == 1:  
                        print(f"X[{p+1}, {w+1}] = 1")

            x_values = []
            y_values = []            
            for p in P:
                for w in W:
                    if X[p, w].solution_value() == 1:  # Only include assignments where X[p, w] is 1
                        x_values.append(w)  # Worker on x-axis
                        y_values.append(p)  # Project on y-axis

            # Create the scatter plot
            plt.figure(figsize=(10, 6))
            plt.scatter(x_values, y_values, color='blue', marker='o')

            # Set labels and title
            plt.xlabel('Workers')
            plt.ylabel('Projects')
            plt.title('Project Assignment to Worker')

            # Set x-axis and y-axis ticks
            plt.xticks(W, [f'Worker {w+1}' for w in W])
            plt.yticks(P, [f'Project {p+1}' for p in P])

            # Show the plot            
            plt.tight_layout()
            plt.show()
                                                    # Gantt Chart

            adjusted_ST_dates = {p: current_date + datetime.timedelta(days=adjusted_ST[p]) for p in P}
            adjusted_FT_dates = {p: current_date + datetime.timedelta(days=adjusted_FT[p]) for p in P}

            gantt_data = []
            for p in P:
                gantt_data.append({
                    'Project': p,
                    'Start': adjusted_ST_dates[p],
                    'Finish': adjusted_FT_dates[p]
                })

            # Convert to DataFrame for plotting
            df = pd.DataFrame(gantt_data)

            # Create Gantt chart
            fig, ax = plt.subplots(figsize=(10, 5))

            # Plot each project as a horizontal bar
            for index, row in df.iterrows():
                ax.barh(row['Project'], (row['Finish'] - row['Start']).days, left=row['Start'], align='center')

            # Format the x-axis as dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=5))
            plt.xticks(rotation=45)

            # Set y-axis as sequential numbers
            ax.set_yticks(range(1, len(P)+1))  # Sequential numbering from 1 to the number of projects
            ax.set_ylabel('Project No')

            # Labels and title
            ax.set_xlabel('Time')
            ax.set_title('Optimized Gantt Chart')

            plt.tight_layout()
            plt.show()
            
            return redirect('dashboard')


    return render(request, 'dashboard/upload.html')  # Render the form again if not POST

